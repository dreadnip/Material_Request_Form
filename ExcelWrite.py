# ------------------------------------------------------------------------ #
# Title: Excel Write
# Description: This program writes the MRF data to an excel file

# ------------------------------------------------------------------------ #
from openpyxl import load_workbook
import datetime
import pickle

if __name__ == "__main__":
    raise Exception("This file is not meant to ran by itself")


# open work book
# todo: create function?
path = r"C:\Users\alcom\Documents\otv_v2\gitignore\MRF_template.xlsx"
wb = load_workbook(path)
ws = wb.active


# create a function with item attribute inputs and write the input to the excel template
# todo: create index variable to write the index to the document.
class XL:

    @staticmethod
    def xl_write(pickle_filename, savedir, vendor):
        with open(pickle_filename, 'rb') as f:
            input_list = pickle.load(f)
        column_letters = ['D', 'E', 'F', 'H', 'I', 'J']
        z = 19
        for item in input_list:
            z += 1
            # todo: figure out why this is only printing the characters of the list not the list items
            for k, letter in enumerate(column_letters):
                ws[letter + str(z)] = item[k]
        d = datetime.datetime.today()
        savename = d.strftime('%m%d%y__%M') + "--ASC " + str(vendor)
        savepath = savedir + '\\' + savename + ".xlsx"
        wb.save(savepath)
        return print("File was saved to" + savepath + "as " + savename)

    # # if/else statements for service level
    # if 'STANDARD' in service_lvl.upper():
    #     ws['F16'] = "CAL NO DATA."
    # elif 'ADVANCED' in service_lvl.upper():
    #     ws['F16'] = "CAL W/ AF-AL DATA"
    # elif 'PREMIER' in service_lvl.upper():
    #     ws['F16'] = "CALIBRATE TO ISO170025 ACCREDITATION"
    # elif custom_lvl:
    #     ws['F16'] = input("What Service level does the item require?")

    # function to save the file with appropriate naming schema
    # todo: add initials variable to allow for multiple users initials
    # todo: create if/else which can navigate the file directory to save
    #  the file automatically to the correct month/year folder

    @staticmethod
    def xlsave(savedir, vendor):
        """
        This function will save the new MRF to the appropriate directory. Save name is "MMDDYYhour-ASC *Vendor*"
        :param savedir: file directory where the excel file will be saved
        :param vendor: Vendor to which the item will be sent to
        :return: String statement confirming the excel document has been saved.
        """

        d = datetime.datetime.today()
        savename = d.strftime('%m%d%y%M') + "-ASC " + str(vendor)
        savepath = savedir + savename + ".xlsx"
        wb.save(savepath)
        return "File was saved to" + savepath + "as " + savename
