# ------------------------------------------------------------------------ #
# Title: Excel Write
# Description: This program writes the MRF data to an excel file

# ------------------------------------------------------------------------ #
from openpyxl import load_workbook
import datetime
import os
import pickle

if __name__ == "__main__":
    raise Exception("This file is not meant to ran by itself")

# open work book
template_path = os.getcwd() + "\\MRF_template.xlsx"
wb = load_workbook(template_path)
ws = wb.active

# define date constant
d = datetime.datetime.today()

# create a function with item attribute inputs and write the input to the excel template
class XL:

    @staticmethod
    def write_item(item_list):
        """Writes Item data to the excel sheet.
        :param item_list: list passed from the new_item instantiation in the form of .mrf_list_format
        :return don't return anything just do it"""

        column_letters = ['A', 'D', 'E', 'F', 'H', 'I']
        # The first row of listing the items begins on 20
        # fills out the rows
        z = 19
        for item in item_list:
            z += 1
            for k, letter in enumerate(column_letters):
                ws[letter + str(z)] = item[k]
    @staticmethod
    def save_sheet(savedir, vendor, initials):
        """This function will save the new MRF to the appropriate directory.
        Save name is "MMDDYY_hour_minute-initials vendor"
        :param savedir: file directory where the excel file will be saved
        :param vendor: Vendor to which the item will be sent to
        :param initials: users initials
        :return: String statement confirming the excel document has been saved, with which path its been saved to
        """

        # savename is mm/dd/yy__hour.minute--Initials Vendor name
        savename = d.strftime('%m%d%y_%H_%M') + "-" + initials + ' ' + str(vendor)
        savepath = savedir + '\\' + savename + ".xlsx"
        wb.save(savepath)
        return print("File was saved to: " + savedir + "\nFilename: " + savename)

    # function to save the file with appropriate naming schema
    # todo: add initials variable to allow for multiple users initials
    # todo: create if/else which can navigate the file directory to save
    #       the file automatically to the correct month/year folder on the work computers
    @staticmethod
    def write_project(index, gageid, company):
        """writes the company name and gageid cells."""
        ws['D6'] = company
        # if there are more than one gageids being passed, write "multiple"
        if index > 0:
            ws['D7'] = 'MULTIPLE'
        else:
            ws['D7'] = str(gageid)
    # xl.write_mrf_details(vendor, requestor, RMA, shipping info, date requested, date required, cert_type, interval, cal to mfg specs)

    @staticmethod
    # def write_mrf_details(initials, vendor, requestor, RMA, shipping_info,  cert_type, interval, cal_specs):
    def write_mrf_details(initials):
        # material_request_num
        ws['J3'] = d.strftime('%m%d%y_%H_%M') + "-" + initials
        # date requested, date required
        ws['D14'] = d.strftime('%m/%d/%Y')
        # adds one month to the date
        ws['D15'] = str((int(d.strftime('%m')) + 1)) + d.strftime('/%d/%Y')

        # vendor.get_address()
        # requestor = User.full_name() -- use pickling to save the user info?
        ws['F11'] = 'Alex Compeau'
        # cert_type = Item.cert_type() unless otherwise specified
        # interval = Item.interval()
        # cal_specs = "cal to mfg specs" unless otherwise specified.


